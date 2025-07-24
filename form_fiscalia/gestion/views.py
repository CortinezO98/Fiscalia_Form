from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Q
from django.views.decorators.http import require_GET
from django.utils.timezone import make_aware, localtime
from usuarios.views import ValidarRolUsuario, en_grupo
from usuarios.enums import Roles
from .models import *
from .utils import RegistrarError  # ← AÑADIR ESTA LÍNEA
from io import BytesIO
from datetime import date, datetime
import openpyxl
from openpyxl.utils import get_column_letter
import inspect

def index(request):
    if request.user.is_authenticated:
        if ValidarRolUsuario(request, Roles.ADMINISTRADOR.value):
            return redirect('dashboard_admin')
        elif ValidarRolUsuario(request, Roles.SUPERVISOR.value):
            return redirect('dashboard_supervisor')
        elif ValidarRolUsuario(request, Roles.AGENTE.value):
            return redirect('dashboard_agente')
        elif ValidarRolUsuario(request, Roles.ABOGADO.value):  # NUEVA REDIRECCIÓN
            return redirect('dashboard_abogado')
        
        messages.warning(request, "Usted no esta autorizado.")
        return redirect('logout')
    else:
        return redirect('login')


@login_required
@en_grupo([Roles.ADMINISTRADOR.value, Roles.SUPERVISOR.value, Roles.AGENTE.value])
def crear_evaluacion(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                cid = request.POST.get('cuidadano_id')
                if cid:
                    ciudadano = Ciudadano.objects.get(id=cid)
                    # Actualizar campos
                    ciudadano.tipo_identificacion_id = request.POST['tipo_identificacion']
                    ciudadano.numero_identificacion = request.POST['numero_identificacion']
                    ciudadano.nombre = request.POST['nombre']
                    ciudadano.correo = request.POST.get('correo', '')
                    ciudadano.telefono = request.POST.get('telefono', '')
                    ciudadano.direccion_residencia = request.POST.get('direccion_residencia', '')
                    ciudadano.pais_id = request.POST.get('pais') or None
                    ciudadano.ciudad = request.POST.get('ciudad', '')
                    ciudadano.save()
                else:
                    ciudadano = Ciudadano.objects.create(
                        tipo_identificacion_id=request.POST['tipo_identificacion'],
                        numero_identificacion=request.POST['numero_identificacion'],
                        nombre=request.POST['nombre'],
                        correo=request.POST.get('correo', ''),
                        telefono=request.POST.get('telefono', ''),
                        direccion_residencia=request.POST.get('direccion_residencia', ''),
                        pais_id=request.POST.get('pais') or None,
                        ciudad=request.POST.get('ciudad', '')
                    )

                # Crear la evaluación con la nueva estructura
                categoria_id = request.POST.get('subcategoria') or request.POST.get('categoria')
                
                # Si categoria_id es "0", significa que es una tipificación sin categorías
                if categoria_id == "0":
                    categoria_id = None

                # Procesar "Se comunica de una URI"
                se_comunica_uri = request.POST.get('se_comunica_uri')
                se_comunica_uri_bool = None
                if se_comunica_uri == '1':
                    se_comunica_uri_bool = True
                elif se_comunica_uri == '0':
                    se_comunica_uri_bool = False
                
                ciudad_municipio_uri = ''
                if se_comunica_uri_bool:
                    ciudad_municipio_uri = request.POST.get('ciudad_municipio_uri', '')

                # Procesar "Consulta jurídica"
                consulta_juridica = request.POST.get('consulta_juridica')
                consulta_juridica_bool = None
                if consulta_juridica == '1':
                    consulta_juridica_bool = True
                elif consulta_juridica == '0':
                    consulta_juridica_bool = False
                
                abogado_id = None
                if consulta_juridica_bool:
                    abogado_id = request.POST.get('abogado') or None        

                # CREAR LA EVALUACIÓN PRIMERO
                evaluacion = Evaluacion.objects.create(
                    conversacion_id=request.POST['conversacion_id'],
                    observacion=request.POST['observacion'],
                    ciudadano=ciudadano,
                    user=request.user,
                    # Nuevos campos
                    tipo_canal_id=request.POST.get('tipo_canal'),
                    segmento_id=request.POST.get('segmento'),
                    segmento_ii_id=request.POST.get('segmento_ii') or None,
                    tipificacion_id=request.POST.get('tipificacion'),
                    # Campo tradicional - CORREGIDO
                    categoria_id=categoria_id,  # ← Usar la variable ya procesada
                    
                    # Campo especial para otros delitos
                    cual_otro_delito=request.POST.get('cual_otro_delito', ''),
                    se_comunica_uri=se_comunica_uri_bool,
                    ciudad_municipio_uri=ciudad_municipio_uri,
                    consulta_juridica=consulta_juridica_bool,
                    abogado_id=abogado_id
                )

                # ==================== ASIGNACIÓN AUTOMÁTICA A ABOGADO ====================
                if consulta_juridica_bool and abogado_id:
                    try:
                        abogado = Abogado.objects.get(id=abogado_id)
                        
                        # Determinar prioridad basada en la tipificación o categoría
                        prioridad = determinar_prioridad_caso(evaluacion)
                        
                        # Crear el caso para el abogado
                        caso_abogado = CasoAbogado.objects.create(
                            evaluacion=evaluacion,
                            abogado=abogado,
                            prioridad=prioridad,
                            estado='PENDIENTE'
                        )
                        
                        messages.success(
                            request, 
                            f"Evaluación guardada correctamente. Caso asignado automáticamente al abogado {abogado.nombre}."
                        )
                        
                        # TODO: Aquí podrías agregar notificación por email al abogado
                        # enviar_notificacion_abogado(caso_abogado)
                        
                    except Abogado.DoesNotExist:
                        messages.warning(request, "Evaluación guardada, pero el abogado seleccionado no existe.")
                else:
                    messages.success(request, "Evaluación guardada correctamente.")

        except Exception as e:
            RegistrarError(inspect.currentframe().f_code.co_name, str(e), request)
            messages.error(request, "Ocurrió un error al guardar la evaluación.")
        
        return redirect('index')

    tiposIdentificacion = TipoIdentificacion.objects.all()
    tipos_canal = TipoCanal.objects.all()
    paises = Pais.objects.all()
    abogados = Abogado.objects.filter(activo=True).order_by('nombre')

    if not tiposIdentificacion or not tipos_canal:
        messages.warning(request, "En este momento no es posible generar una tipificación. Por favor, contacte a soporte.")
        return redirect('index')

    return render(request, 'usuarios/evaluaciones/crear_evaluacion.html', {
        'tiposIdentificacion': tiposIdentificacion,
        'tipos_canal': tipos_canal,
        'paises': paises,
        'abogados': abogados,
    })

# ==================== FUNCIÓN AUXILIAR PARA DETERMINAR PRIORIDAD ====================
def determinar_prioridad_caso(evaluacion):
    """
    Determina la prioridad del caso basado en la tipificación o categoría
    """
    try:
        categoria_nombre = evaluacion.categoria.nombre.upper()
        tipificacion_nombre = evaluacion.categoria.tipificacion.nombre.upper() if evaluacion.categoria.tipificacion else ""
        
        # Casos de alta prioridad
        palabras_alta_prioridad = [
            'VIOLENCIA', 'SEXUAL', 'INTRAFAMILIAR', 'MENOR', 'INFANTIL', 
            'HOMICIDIO', 'SECUESTRO', 'EXTORSIÓN', 'AMENAZA', 'URGENTE'
        ]
        
        # Casos de baja prioridad  
        palabras_baja_prioridad = [
            'CONSULTA', 'INFORMACIÓN', 'ORIENTACIÓN', 'DOCUMENTOS', 'COPIA'
        ]
        
        texto_completo = f"{categoria_nombre} {tipificacion_nombre}"
        
        for palabra in palabras_alta_prioridad:
            if palabra in texto_completo:
                return 'ALTA'
                
        for palabra in palabras_baja_prioridad:
            if palabra in texto_completo:
                return 'BAJA'
                
        return 'MEDIA'  # Prioridad por defecto
        
    except Exception:
        return 'MEDIA'  # En caso de error, asignar prioridad media


@login_required
@en_grupo([Roles.ADMINISTRADOR.value, Roles.SUPERVISOR.value, Roles.AGENTE.value])
def buscar_tipificacion(request):
    """
    Busca Evaluacion(s) únicamente por numero_identificacion (ciudadano),
    paginadas de a 10. Devuelve todos los campos relacionados.
    """
    query = request.GET.get('q', '').strip()
    evaluaciones = Evaluacion.objects.none()

    if query:
        qs = (
            Evaluacion.objects
            .select_related(
                'ciudadano__tipo_identificacion',
                'ciudadano__pais',
                'categoria__tipificacion',
                'categoria__categoria_padre',
                'user',
                'tipo_canal',
                'segmento',
                'segmento_ii',
                'tipificacion'
            )
            .filter(ciudadano__numero_identificacion__icontains=query)
            .order_by('-fecha')
        )

        paginator = Paginator(qs, 10)
        page_number   = request.GET.get('page') or 1
        evaluaciones  = paginator.get_page(page_number)
    else:
        paginator = None

    return render(request, 'usuarios/evaluaciones/buscar_tipificacion.html', {
        'evaluaciones': evaluaciones,
        'paginator':    paginator,
        'page_obj':     evaluaciones,
        'is_paginated': evaluaciones.has_other_pages() if evaluaciones else False,
        'query':        query,
    })


@login_required
@en_grupo([Roles.ADMINISTRADOR.value, Roles.SUPERVISOR.value, Roles.AGENTE.value])
def reportes_view(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin    = request.GET.get('fecha_fin')

    qs = (
        Evaluacion.objects
        .select_related(
            'ciudadano__tipo_identificacion',
            'ciudadano__pais',
            'categoria__tipificacion',
            'categoria__categoria_padre',
            'user',
            'tipo_canal',
            'segmento',
            'segmento_ii',
            'tipificacion'
        )
        .order_by('-fecha')
    )

    if fecha_inicio and fecha_fin:
        try:
            fmt = '%Y-%m-%d'
            fi = make_aware(datetime.strptime(fecha_inicio, fmt))
            ff = make_aware(datetime.strptime(fecha_fin,    fmt))
            qs = qs.filter(fecha__range=(fi, ff))
        except ValueError:
            pass

    paginator = Paginator(qs, 10)
    page_obj  = paginator.get_page(request.GET.get('page') or 1)

    return render(request, 'usuarios/reportes.html', {
        'reportes':     page_obj,
        'paginator':    paginator,
        'page_obj':     page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'hoy':          date.today().isoformat(),
    })


@require_GET
@login_required
@en_grupo([Roles.ADMINISTRADOR.value, Roles.SUPERVISOR.value])
def exportar_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin    = request.GET.get('fecha_fin')

    qs = (
        Evaluacion.objects
        .select_related(
            'ciudadano__tipo_identificacion',
            'ciudadano__pais',
            'categoria__tipificacion',
            'categoria__categoria_padre',
            'user',
            'tipo_canal',
            'segmento',
            'segmento_ii',
            'tipificacion',
            'abogado'
        )
        .order_by('-fecha')
    )

    if fecha_inicio and fecha_fin:
        try:
            fmt = '%Y-%m-%d'
            fi = make_aware(datetime.strptime(fecha_inicio, fmt))
            ff = make_aware(datetime.strptime(fecha_fin,    fmt))
            qs = qs.filter(fecha__range=(fi, ff))
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reportes"
    headers = [
        'Fecha', 'Tipo Documento', 'Número ID', 'Nombre',
        'Correo', 'Teléfono', 'País', 'Ciudad', 'Dirección',     
        'ID Conversación', 'Tipo Canal', 'Segmento', 'Segmento II',
        'Tipificación', 'Categoría', 'Categoría Padre', 'Cuál Otro Delito',
        'Observación', 'Se comunica URI', 'Ciudad/Municipio URI', 'Consulta Jurídica',
        'Delito Código', 'Delito Nombre', 'Interacción Directa', 'Habeas Corpus',
        'Tutela', 'Observaciones Abogado', 'Abogado', 'Usuario'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, ev in enumerate(qs.iterator(), start=2):
        ciu = ev.ciudadano
        se_comunica_uri = 'SÍ' if ev.se_comunica_uri is True else 'NO' if ev.se_comunica_uri is False else 'N/A'
        consulta_juridica = 'SÍ' if ev.consulta_juridica is True else 'NO' if ev.consulta_juridica is False else 'N/A'
        abogado_nombre = ev.abogado.nombre if ev.abogado else ''

        caso_abogado = getattr(ev, 'caso_abogado', None)
        delito_codigo = caso_abogado.delito.codigo if caso_abogado and caso_abogado.delito else ''
        delito_nombre = caso_abogado.delito.nombre if caso_abogado and caso_abogado.delito else ''
        interaccion_directa = 'SÍ' if caso_abogado and caso_abogado.interaccion_directa_usuario is True else 'NO' if caso_abogado and caso_abogado.interaccion_directa_usuario is False else 'N/A'
        habeas_corpus = 'SÍ' if caso_abogado and caso_abogado.habeas_corpus is True else 'NO' if caso_abogado and caso_abogado.habeas_corpus is False else 'N/A'
        tutela = 'SÍ' if caso_abogado and caso_abogado.tutela is True else 'NO' if caso_abogado and caso_abogado.tutela is False else 'N/A'
        observaciones_abogado = caso_abogado.observaciones_abogado if caso_abogado else ''
        data = [
            localtime(ev.fecha).strftime('%Y-%m-%d %H:%M'),
            ciu.tipo_identificacion.nombre,
            ciu.numero_identificacion,
            ciu.nombre,
            ciu.correo or '',
            ciu.telefono or '',
            ciu.pais.nombre if ciu.pais else '',
            ciu.ciudad or '',
            ciu.direccion_residencia or '',
            ev.conversacion_id,
            # Nuevos campos
            ev.tipo_canal.nombre if ev.tipo_canal else '',
            ev.segmento.nombre if ev.segmento else '',
            ev.segmento_ii.nombre if ev.segmento_ii else '',
            ev.tipificacion.nombre if ev.tipificacion else '',
            ev.categoria.nombre if ev.categoria else 'Sin categoría específica',
            ev.categoria.categoria_padre.nombre if ev.categoria and ev.categoria.categoria_padre else '',
            ev.cual_otro_delito or '',
            ev.observacion,
            se_comunica_uri,
            ev.ciudad_municipio_uri or '',
            consulta_juridica,
            delito_codigo,
            delito_nombre,
            interaccion_directa,
            habeas_corpus,
            tutela,
            observaciones_abogado, 
            abogado_nombre,
            ev.user.username
        ]
        for col, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col, value=value)

    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].auto_size = True

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resp = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="reportes_tipificaciones.xlsx"'
    return resp